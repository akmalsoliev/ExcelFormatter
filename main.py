import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

import numpy as np 


class Excel_Fomatter:
    def __init__(self, file_path:str, sheet:str, index_size:int=1, column_size:int=1):
        """
        Produces formated Excel file.
        :param file_path: Full path of the excel file.
        :param sheet: Sheet that will be formatted.
        :param index_size: if working with multilayer index then specify the size of it (how many columns is the multilayer index).
        :param column_size: if working with multilayer column then specify the size of it (how many rows is the multilayer column).
        """
        self.file_path = file_path
        self.sheet = sheet
        self.index_size = index_size
        self.column_size = column_size
        self.workbook = self.__get_workbook__()
        self.worksheet = self.__get_worksheet__()
        self.xl_cols, self.xl_rows = self.__get_shape__()

    def __get_workbook__(self):
        return openpyxl.load_workbook(self.file_path)

    def __get_worksheet__(self):
        return self.workbook[self.sheet]

    def __get_shape__(self):
        xl_rows, xl_cols = np.array([*self.worksheet.columns]).shape
        return xl_rows, xl_cols

    def del_row(self, row, column):
        self.worksheet.delete_rows(row, column)
        return "Row deletion complete!"

    def set_column_percentage(self, start_col:int, start_row:int, steps:int, skip_rows:int=1):
        """
        Formats columns to percentage
        :param start_col: Starting column.
        :param steps: Every X column will be formatted as percentage.
        :param skip_rows: Number of rows that should be skipped in the formatting (number of header rows).
        """
        for column in np.arange(start_col, self.xl_cols+1, steps):
            for row in np.arange(start_row, self.xl_rows):
                if row > skip_rows:
                    cell = self.worksheet.cell(row=row, column=column)
                    setattr(cell, "style", "Percent")
        return "Percentage formatting is complete!"

    def set_row_percentage(self, start_col:int, start_row:int, num_rows:int, steps:int=1,skip_cols:int=1):
        """
        Formats row to percentage
        :param start_row: Starting row.
        :param steps: Every X column will be formatted as percentage.
        :param num_rows: Number of rows that should be formatted starting from start_row.
        """
        for rows in range(start_row,start_row+num_rows,steps):
            for cols in np.arange(start_col, self.xl_cols+1):
                if cols > skip_cols:
                    cell = self.worksheet.cell(row=rows, column=cols)
                    setattr(cell, "style", "Percent")
        return "Percentage formatting is complete!"

    def set_row_currency(self, start_col:int, start_row:int, num_rows:int, steps:int=1, skip_cols:int=1):
        """
        Formats row to currency
        :param start_row: Starting row.
        :param steps: Every X column will be formatted as currency $ .
        :param num_rows: Number of rows that should be formatted starting from start_row.
        """
        
        for rows in range(start_row,start_row+num_rows,steps):
            for cols in np.arange(start_col, self.xl_cols+1):
                if cols > skip_cols:
                    cell = self.worksheet.cell(row=rows, column=cols)
                    setattr(cell, "number_format", u'$#,##0_-')
        return "Currency formatting is complete!"

    def set_row_percentage_dec(self, start_col:int, start_row:int, num_rows:int, steps:int=1,skip_cols:int=1,decimal:int=1):
        """
        Formats row to percentage. comma
        :param start_row: Starting row.
        :param steps: Every X column will be formatted as percentage.
        :param num_rows: Number of rows that should be formatted starting from start_row.
        :param decimal: Number of decimal.
        """

        for rows in range(start_row,start_row+num_rows,steps):
            for cols in np.arange(start_col, self.xl_cols+1):
                if cols > skip_cols:
                    cell = self.worksheet.cell(row=rows, column=cols)
                    if decimal==0:
                        cell.number_format = "0%"
                    else:
                        cell.number_format = "0."+ decimal* "0"+"%"
        return "Percentage formatting is complete!"

    def set_row_number_dec(self, start_col:int, start_row:int, num_rows:int, steps:int=1,skip_cols:int=1,decimal:int=1):
        """
        Formats row to number.comma
        :param start_row: Starting row.
        :param steps: Every X column will be formatted as percentage.
        :param num_rows: Number of rows that should be formatted starting from start_row.
        :param decimal: Number of decimal.
        """

        for rows in range(start_row,start_row+num_rows,steps):
            for cols in np.arange(start_col, self.xl_cols+1):
                if cols > skip_cols:
                    cell = self.worksheet.cell(row=rows, column=cols)
                    if decimal==0:
                        cell.number_format = '0'
                    else:
                        cell.number_format = "0."+ decimal* "0"
                        
        return "Percentage formatting is complete!"
        
    def freeze_panes(self, cell:str=None, row:int=None, column:int=None):
        if cell and (row or column):
            raise """Need to specify either `cell` or `row` and `column`. Cannot declare `cell` and `row` or `column`."""
        elif cell:
            self.worksheet.freeze_panes = cell
            return "Pane freeze is complete"
        elif row and column:
            ws = self.worksheet
            selected_cell = ws.cell(row, column)
            self.worksheet.freeze_panes = selected_cell
            return "Pane freeze is complete"
        else:
            raise "Need to specify either `cell` or `row` and `column`"

    def __fill_color__(self, color:str):
            return PatternFill(start_color=color, end_color=color, fill_type='solid')

    def color_columns(self, start:int, steps:int, color0:str="FFFFFF", color1:str="EEEBFF"):
        # Color Fill
        col_fill0 = self.__fill_color__(color0)
        col_fill1 = self.__fill_color__(color1)

        start_col_range = np.arange(start, self.xl_cols, steps)
        end_col_range = np.arange(start+steps, self.xl_cols+steps, steps)
        cols = [x for x in self.worksheet.columns]

        for i, (start, end) in enumerate(zip(start_col_range, end_col_range)):
            for cell in np.array(cols[start:end]).flatten():
                if i%2:
                    cell.fill = col_fill0
                else:
                    cell.fill = col_fill1
        return "Column coloring is complete"

    def header_color(self, color:str="B4C6E7", header_num_rows:int=2):
        head_fill = self.__fill_color__(color)
        try:
            for row in self.worksheet["1:{}".format(header_num_rows)]:
                for cell in row:
                    cell.fill = head_fill
        except TypeError:
            for cell in self.worksheet[header_num_rows]:
                cell.fill = head_fill
        return "Header coloring is complete"

    def color_row(self, start_row:int, end_row:int, color:str="FCD5B4"):
        color_fill = self.__fill_color__(color)
        if start_row - end_row == 0:
            for cell in self.worksheet["{}:{}".format(start_row, end_row)]:
                cell.fill = color_fill
        else:
            for row in self.worksheet["{}:{}".format(start_row, end_row)]:
                for cell in row:
                    cell.fill = color_fill
        return "Row coloring is complete"

    def format_border(self):
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        [setattr(cell, "border", border) for cell in np.array([*self.worksheet.columns]).flatten()]
        return "Border formatting is complete"

    def columns_width(self, padding:int=3):

        dim_holder = DimensionHolder(worksheet=self.worksheet)

        for col in np.arange(1, self.xl_cols + 1):
            # intiating the width of the column as 0, as it will serve base size.
            width = 0
            for row in np.arange(1, self.xl_rows + 1):
                cell_value = self.worksheet.cell(column=col, row=row).value
                if cell_value:
                    # finding what's the len of the cell's values and then setting it
                    # as a new base `width`
                    cell_len = len(str(cell_value))
                    if cell_len > width:
                        width = cell_len + padding

            dim_holder[get_column_letter(col)] = (
                ColumnDimension(
                    self.worksheet,
                    min=col,
                    max=col,
                    width=width)
            )

        self.worksheet.column_dimensions = dim_holder
        return "Column width adjustment is complete"

    def merged_header_alternate_fill(self, start_cell=(3,2), color1="6A8ED0",color2="9EB2DE"):
        """
        Formats header alternated by merged cell
        :param start_cell: Starting cell as list
        :param color1 & color2: define which color to alternate
        """ 
        ws=self.worksheet
        m_c_range=ws.merged_cells.ranges
        #ROW OF FIRST LEVEL INDEXING : YEAR_MONTH
        #  row_range=3
        start_row=start_cell[0] 

        #COLUN START OF HEADER
        start_col=start_cell[1] 

        #list of cell to merge
        #output of first loop to identify which cell to fill
        #method .fill takes the first cell top right to fill
        merged_cell_complete=[]
        
        #index for list merged_cell_complete
        j, f = -1, 0
        #BUILD INDEX FOR HEADER, INCLUDING NON MERGED CELLS
        #RETURNS LEFT INDEX OF CELL IF IN A MERGE RANGE OR THE SAME CELL IF NOT
        for i in range(len(m_c_range)):
            m_c_left=m_c_range[i].left
            m_c_right=m_c_range[i].right
            #first time of loop starts from first column that could not be included in first merged cell
            #insert column until the first merged cell then take the merged cell
            if i==0:
                s=[(start_row,start_col)]
                while s!=m_c_left:
                    merged_cell_complete.append(s)
                    j=j+1
                    f=f+1
                    s=[(start_row,start_col+f)]
                merged_cell_complete.append(m_c_left)
                j=j+1
            else:
                new_cell=m_c_range[i-1].right
                new_cell=[(new_cell[0][0],new_cell[0][1]+1)]
                while new_cell!=m_c_left:
                    merged_cell_complete.append(new_cell)
                    j=j+1
                    new_cell=[(new_cell[0][0],new_cell[0][1]+1)]
                merged_cell_complete.append(m_c_left)
                j=j+1
            #stop condition if non-merged cell to be filled take place at the end
            if i==len(m_c_range)-1:
                ii=m_c_right[0][1]+1
                if ws.cell(row=start_row,column=ii).value!=None:
                    while ws.cell(row=start_row,column=ii).value!=None:
                        new_cell=[(start_row,ii)]
                        merged_cell_complete.append(new_cell)
                        ii=ii+1
                        j=j+1
        #ALTERNATE FILL (HEADER)
        foo=True
        for m_c_list in merged_cell_complete:
            cell_index=m_c_list[0]
            if foo==True:
                ws.cell(row=cell_index[0],column=cell_index[1]).fill=PatternFill("solid", start_color=color1)
                foo=False
            else:
                ws.cell(row=cell_index[0],column=cell_index[1]).fill=PatternFill("solid", start_color=color2)
                foo=True
        print('MERGED CELL FILLED')

    def save(self):
        self.workbook.save(self.file_path)
        return "File saving is complete"
